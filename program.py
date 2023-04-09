import sys
import os
import glob
from dataclasses import dataclass
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from datetime import datetime
from tkinter import filedialog
from tkinter import *

@dataclass
class Student:
    name: str 
    time_total: float

@dataclass
class SubdivisionRecord:
    counter: int
    time_total: float
    students: []

def readData(filename):
    book = load_workbook(filename=filename, data_only=True)
    book.active = 0
    sheet = book.active
    groups = dict()
    # get date from file
    date = sheet['B'+str(1)].value[40:]
    # calculate number of days
    i = 5
    running = True
    # number_of_weekdays is excluding saturdays and sundays
    number_of_weekdays = 0
    number_of_days = 0
    while(running):
        if sheet['B' + str(i)].value and 'Суббота' not in sheet['B' + str(i)].value and 'Воскресенье' not in sheet['B' + str(i)].value:
            number_of_weekdays += 1
        number_of_days += 1
        i += 1
        if(sheet['B'+str(i)].value == None):
            running = False
    number_of_weekdays -= 1
    number_of_days -= 1
    i = 5
    running = True
    while(i < sheet.max_row and running):
        # get name and subdivision
        name = sheet['B'+str(i)].value
        subdivision = sheet['C'+str(i)].value
        # create subdivision attendance array
        if not subdivision in groups:
            groups[subdivision] = SubdivisionRecord(0, 0.0, [])
        groups[subdivision].counter += 1
        # skip days cells
        i += number_of_days + 1
        # calculate total tame
        hours, minutes = sheet['G'+str(i)].value.split(':')
        students_time_total = int(hours) + int(minutes) / 60
        groups[subdivision].time_total += students_time_total
        groups[subdivision].students.append(Student(name, students_time_total))
        i += 1
        # out of bounds check
        if(sheet['C'+str(i)].value == None):
            running = False

    return (date, number_of_weekdays, groups)

def writeData(filename, newfilename, date, number_of_weekdays, groups):
    book = Workbook()
    # First sheet
    sheet = book.active
    sheet.title = 'Лист 1'
    myFont = Font(name='Calibri', color='FF0000')
    row = 2
    column = 1
    sheet.column_dimensions['A'].width = 35
    sheet.column_dimensions['B'].width = 35
    cell = sheet.cell(row=row, column=column)
    cell.value = "Группа"
    cell = sheet.cell(row=row, column=column+1)
    cell.value = "Количество студентов"
    cell = sheet.cell(row=row+1, column=column)
    cell.value = "Ожид. / Реал. часов"
    cell = sheet.cell(row=row+1, column=column+1)
    cell.value = "Процент посещаемости"
    cell = sheet.cell(row=row+2, column=column)
    cell.value = "Ф.И.О. меньше {percent}% посещаемости".format(percent=percent_of_attendance)
    cell = sheet.cell(row=row+2, column=column+1)
    cell.value = "Часы"
    row = 2
    column = 4
    for group in groups:
        cell = sheet.cell(row=row, column=column)
        cell.value = group
        sheet.column_dimensions[get_column_letter(cell.column)].width = 35
        cell = sheet.cell(row=row, column=column+1)
        cell.value = groups[group].counter
        cell = sheet.cell(row=row+1, column=column)
        cell.value = "{expected} / {actual}".format(expected=round(groups[group].counter * number_of_weekdays * 3.2), actual=round(groups[group].time_total))
        cell = sheet.cell(row=row+1, column=column+1)
        cell.value = "{value}%".format(value=round(groups[group].time_total * 100.0 / (groups[group].counter * number_of_weekdays * 3.2)))
        row += 2

        for student in sorted(groups[group].students, key=lambda s: s.name):
            if (student.time_total / number_of_weekdays) <= (3.2/100.0*int(percent_of_attendance)):
                cell = sheet.cell(row=row, column=column)
                cell.value = student.name

                cell = sheet.cell(row=row, column=column+1)
                cell.value = "{time:.1f}".format(time=student.time_total)
                if student.time_total == 0.0:
                    cell.font = myFont
                row += 1
        column += 1

        row = 2
        column += 2
    # Second sheet
    sheet = book.create_sheet('Лист 2')
    row = 1
    column = 1
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 12
    cell = sheet.cell(row=row, column=column)
    cell.value = "Группа"
    cell = sheet.cell(row=row, column=column+1)
    cell.value = "Кол-во студ."
    cell = sheet.cell(row=row, column=column+2)
    cell.value = "Ожид. часов"
    cell = sheet.cell(row=row, column=column+3)
    cell.value = "Реал. часов"
    cell = sheet.cell(row=row, column=column+4)
    cell.value = "%"
    row += 1

    for group in groups:
        cell = sheet.cell(row=row, column=column)
        cell.value = group
        cell = sheet.cell(row=row, column=column+1)
        cell.value = groups[group].counter
        cell = sheet.cell(row=row, column=column+2)
        cell.value = round(groups[group].counter * number_of_weekdays * 3.2)
        cell = sheet.cell(row=row, column=column+3)
        cell.value = round(groups[group].time_total)
        cell = sheet.cell(row=row, column=column+4)
        cell.value = round(groups[group].time_total * 100.0 / (groups[group].counter * number_of_weekdays * 3.2))
        row += 1
    book.save(newfilename)
        
directory = ""

def select_file():
    global filenames
    filename = filedialog.askopenfilename(initialdir = "./",title = "Выбрать файл",filetypes = (("excel files","*.xlsx"),("all files","*.*")))
    if not filenames:
        filenames.append(filename)
    else:
        filenames[0] = filename


def filter_indirs(indir):
    if '.' in indir:
        return False
    return True

def select_dir():
    global filenames
    global directory
    directory = filedialog.askdirectory(initialdir = "./", title = "Выбрать папку")
    if directory != ():
        newdirname = directory.replace(directory.split('/')[-1], directory.split('/')[-1]+"_посещение")
        indirs = list(filter(filter_indirs, os.listdir(directory)))
        filenames = glob.glob('{}/**/*.xlsx'.format(directory), recursive=True)
        if not os.path.exists(newdirname):
            os.makedirs(newdirname)
        for dir in indirs:
            if not os.path.exists(os.path.join(newdirname, dir)):
                os.makedirs(os.path.join(newdirname, dir))

def start_callback():
    for filename in filenames:
        print(filename)
        date, number_of_weekdays, groups = readData(filename)
        newfilename = ""
        if len(filenames) == 1:
            newfilename = filename[:-5]+"_посещение.xlsx"
        else:
            newfilename = filename.replace(directory.split('/')[-1], directory.split('/')[-1]+"_посещение")
        writeData(filename, newfilename, date, number_of_weekdays, groups)

def percent_selection(v):
    global percent_of_attendance
    percent_of_attendance = v


filenames = []
percent_of_attendance = 0
window = Tk()
window.title("Учет студентов")
filelabel = Label(text="Выберите файл или папку с файлами и нажмите старт")
filelabel.pack()
btn_select_file = Button(window, text="Выбрать файл", command=select_file)
btn_select_file.pack()
btn_select_dir = Button(window, text="Выбрать папку", command=select_dir)
btn_select_dir.pack()
scale = Scale(window, length=250, orient=HORIZONTAL, label="% посещения", from_=0, to=100, command=percent_selection)
scale.pack()
btn_start = Button(window, text="Старт", command=start_callback)
btn_start.pack()
mainloop()
